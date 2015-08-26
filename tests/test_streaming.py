# -*- coding: utf-8 -*-
from collections import namedtuple
import datetime
from io import BytesIO

from django.db import models
from django.test import TestCase

import openpyxl

from rest_framework import serializers as drf_serializers

from rest_framework_oxml_streaming import serializers
from rest_framework_oxml_streaming import streaming

class TestModel(models.Model):
    field_one = models.TextField()
    field_two = models.IntegerField()
    field_three = models.DateTimeField()

class TestSerializer(serializers.OpenXMLSerializer):
    field_one = drf_serializers.CharField()
    field_two = drf_serializers.IntegerField()
    field_three = drf_serializers.ReadOnlyField()
    field_four = serializers.CurrencyField(u'€', verbose_name=u'currency')

class StreamingTestCase(TestCase):
    def test_render(self):
        def test_iterator():
            TestTuple = namedtuple('Test', 'field_one field_two field_three field_four')
            data = [TestTuple('Test string', 2, datetime.datetime(2010, 1, 2, 12, 13, 14), 100)]
            yield TestSerializer(data[0], convert=False).data
            for _i in range(10):
                yield TestSerializer(data, convert=True, many=True).data
        stream = streaming.OpenXMLRenderer().render(
            data=test_iterator(),
            renderer_context={'column_headers': ['field one', 'field two', 'field three', 'field four']}
        )
        stream_bytes = BytesIO()
        for data in stream:
            stream_bytes.write(data)
        wb = openpyxl.load_workbook(stream_bytes)
        self.assertEqual(wb.active.cell('A1').value, 'field one')
        self.assertEqual(wb.active.cell('D1').value, 'field four')
        self.assertEqual(wb.active.cell('A2').value, 'Test string')
        self.assertEqual(wb.active.cell('B3').value, 2)
        self.assertEqual(abs(wb.active.cell('C4').value - datetime.datetime(2010, 1, 2, 12, 13, 14)).seconds, 0)
        self.assertEqual(wb.active.cell('D5').value, 100)
